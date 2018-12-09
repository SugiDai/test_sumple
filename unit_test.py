# coding:utf-8
import openpyxl
import time
import re
from selenium import webdriver
from selenium.webdriver.common.alert import Alert


def init_check(driver, case):
    # imagesフォルダにスクリーンショットを保存
    file_name = "case{}_1.png".format(case['no'])

    driver.save_screenshot("./images/" + file_name)


def set_input(driver, case):

    input_list = case['input_list']
    for input_data in input_list:
        # element = "#{}".format(input_data['id'])
        # driver.find_elements_by_css_selector(selector)
        p_element = driver.find_element_by_id(input_data['id'])
        element = p_element.find_element_by_class_name(input_data['class'])
        element.send_keys(input_data['value'])

    time.sleep(1)

    file_name = "case{}_2.png".format(case['no'])
    driver.save_screenshot("./images/" + file_name)

    # 単一の要素を取得
    driver.find_element_by_id('btn01').click()

    # アラートボタン押下
    Alert(driver).accept()


def checkput(driver, case):

    # imagesフォルダにスクリーンショットを保存
    file_name = "case{}_3.png".format(case['no'])
    driver.save_screenshot("./images/" + file_name)
    check_list = case['check_list']

    for check_data in check_list:
        p_element = driver.find_element_by_id(check_data['id'])
        element = p_element.find_element_by_class_name(check_data['class'])
        value = element.get_attribute("value")
        if(check_data['value']==value):
            print('OK:{}',format(value))
        else:
            print('NG:{}',format(value))


def test_case(case):
    driver = webdriver.Chrome("C:\DRIVERS\chromedriver.exe")
    driver.get("C:\work2\index.html")
    # driver.set_window_size(1250, 1036)
    # driver.execute_script("document.body.style.zoom='90%'")

    init_check(driver, case)

    set_input(driver, case)

    checkput(driver, case)

    # 終了
    driver.close()


def test_phase(case_list):
    print(len(case_list))
    for case in case_list:
        test_case(case)


def get_case_list(sheet):

    count = 0
    calse_list = []
    input_list = []
    check_list = []

    for row in sheet.iter_rows(min_row=6):
        count = count + 1
        if count % 4 == 0:
            case_dict = {'no': count//3, 'input_list': input_list,
                         'check_list': check_list}
            calse_list.append(case_dict)
            input_list = []
            check_list = []

        if row[26].value is not None:
            in_dict = dict()
            in_dict['id'] = str(row[26].value)
            in_dict['class'] = str(row[29].value)
            in_dict['value'] = str(row[33].value)
            input_list.append(in_dict)

        if row[43].value is not None:
            check_dict = dict()
            check_dict['id'] = str(row[43].value)
            check_dict['class'] = str(row[46].value)
            check_dict['value'] = str(row[50].value)
            check_list.append(check_dict)

    case_dict = {'no': count//3, 'input_list': input_list,
                 'check_list': check_list}
    calse_list.append(case_dict)

    return calse_list


def main():
    """
    メイン処理
    """
    # ワークシート読み込み
    wb = openpyxl.load_workbook('unitTestDocument.xlsx')

    for name in wb.sheetnames:

        # 処理対象シート取得
        sheet = wb[name]
        case_list = get_case_list(sheet)
        test_phase(case_list)


if __name__ == "__main__":
    main()
